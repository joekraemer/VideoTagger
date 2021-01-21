import sys
sys.path.append('../')

import logging
import os
import json
import asyncio
from pprint import pprint
import time
import openpyxl as pyxl
import shutil
import boto3
import threading
import progressbar
from pathlib import Path
from botocore.exceptions import ClientError
from concurrent.futures.thread import ThreadPoolExecutor
import requests

from rekognition_objects import (
    RekognitionFace, RekognitionCelebrity, RekognitionLabel,
    RekognitionModerationLabel, RekognitionPerson)

# TODO get this reference to work
# from FindingFilesPackage import filemanager as fm

from rekognition_video_detection import RekognitionVideo

csvTemplate = 'C:/VMShared/GitRepos/PersonalProject/VideoTagger/BlackBoxMetadataTemplate032218.xlsx'
rootDir = 'C:/VMShared/GitRepos/PersonalProject/VideoTagger'
targetDir = 'D:/ResolveTestingOneSecond/Booty'

videoExtensions = ["*.mp4", "*.mov", "*.MP4", "*.avi", "*.mkv", "*.m4v"]

logger = logging.getLogger(__name__)


def FindItemsInDirectory(srcDir, extensionList, recursive=True):
        pathList = []
        for ext in extensionList:
            if recursive:
                ext = "**/" + ext
            pathList.extend(Path(srcDir).glob(ext))
        return pathList

class UploadProgressBar():
    def __init__ (self, size):
        # Create a progress bar class
        self.up_progress = progressbar.progressbar.ProgressBar(maxval=size)

    def upload_progress(self, chunk):
        self.up_progress.update(self.up_progress.currval + chunk)

    def start(self):
        self.up_progress.start()

    def finish(self):
        self.up_progress.finish()

def upload_file(file_path, bucketObj, object_name=None):
    """Upload a file to an S3 bucket

    :param file_path: File to upload
    :param bucketObj: Bucket to upload to
    :param object_name: S3 object name. If not specified then file_name is used
    :return: True if file was uploaded, else False
    """

    # If S3 object_name was not specified, use file_name
    if object_name is None:
        head_tail = os.path.split(file_path)
        object_name = head_tail[1]

    # Create an object
    object = bucketObj.Object(object_name)

    # Get the size of the file 
    statinfo = os.stat(file_path)

    # Create progress bar 
    FileUploadProgress = UploadProgressBar(statinfo.st_size)

    # Upload the file
    try:
        # Start progress bar
        FileUploadProgress.start()

        response = object.upload_file(file_path.__str__(), Callback=FileUploadProgress.upload_progress)
    except ClientError as e:
        logging.error(e)
        return False

    # Stop Progress bar
    FileUploadProgress.finish()

    return object


def CreateNotification(rekogVideoObject):
    name = rekogVideoObject.video_name
    # Can't include . in resource names for sns, sqs
    name_ext = os.path.splitext(name)
    nameString = name_ext[0] + str(time.time_ns())

    # First see if a notification channel already exists under this name
    # there would be a role queue and topic

    # Create a notification channel for this video
    print("Creating notification channel from Amazon Rekognition to Amazon SQS.")
    iam_resource = boto3.resource('iam')
    sns_resource = boto3.resource('sns')
    sqs_resource = boto3.resource('sqs')
    rekogVideoObject.create_notification_channel(
        nameString, iam_resource, sns_resource, sqs_resource)


def OrganizeTagsByConfidence(labelList):
    # takes a list of rekognition_objects.RekognitionLabel

    # this is the lamda function that I will use to organize these rekognition labels
    # Sort by confidence and by name

    def k(i): return (i.confidence, i.name)

    for label in labelList:
        label.to_dict()  # Converts the RekognitionLabel into a dictionary

    labelListSorted = sorted(labelList, key=k, reverse=True)
    return labelListSorted


def RemoveDuplicates(labelList):
    # can't use set with a key so won't work because we have a list of dictionaries
    res = []
    labelsNoDups = []
    for i in labelList:
        if i.name not in res:
            res.append(i.name)
            labelsNoDups.append(i)

    return labelsNoDups


def StripLabelsFromLabelObj(labelList):
    labelListAsString = []
    for labelObj in labelList:
        labelListAsString.append(labelObj.name)
    return labelListAsString

def ParseLabels(labelList):
    # This will likely change as I mess around with additional parameters
    noDups = RemoveDuplicates(labelList)
    organized = OrganizeTagsByConfidence(noDups)
    labelString = StripLabelsFromLabelObj(organized)
    return labelString

def ListToString(lst):
    # Takes a lits of strings, in most of these cases, these will be metadata keywords
    # returns a comma deliniated string of the keywords
    lstStr = ''

    for item in lst:
        lstStr += (str(item))
        lstStr += ', '

    return lstStr


def DetectLabelsWithRekog(path, bckt):
    # Upload each video
    print('Starting ' + path.name + ' upload')
    video_object = upload_file(path, bckt)

    # This is a making a rekognition object
    rekognition_client = boto3.client('rekognition')

    rekog_video = RekognitionVideo.from_bucket(
        video_object, rekognition_client)

    # Create a SNS Notification Channel so that we can get notified when the video analysis is complete
    CreateNotification(rekog_video)

    print("Detecting labels in: " + str(video_object.key))

    labels = rekog_video.do_label_detection()

    labels_parsed = ParseLabels(labels)

    rekog_video.delete_notification_channel()
    return labels_parsed


def GetMetadataFromRekog(path, bckt, BB_CSV_Manager):
    # We want to queue up this function
    md = DetectLabelsWithRekog(path, bckt)

    # Once we get data back, we want to lock the csv so it is only
    # handled by one thread and then update it with the metadata for this video
    BB_CSV_Manager.LockCSV()
    BB_CSV_Manager.AddVideo(path.name, md)
    BB_CSV_Manager.SaveCSV()
    BB_CSV_Manager.UnlockCSV()

def CreateBlackBoxCSVWithRekog(path_list, bckt, csvOb, enableMultithread = True):


    if enableMultithread:
        with ThreadPoolExecutor(max_workers=2) as executor:
            for path in path_list:
                executor.submit(GetMetadataFromRekog, path, bckt, csvOb)
    
    else:
        for path in path_list:
            GetMetadataFromRekog(path, bckt, csvOb)

    return

class BlackBoxCSVManager:
    ######
    # This class is used to open and manipulate CSV files used for automatically adding meta data
    ######

    def __init__(self, filePath=None):

        self.CSVPath = filePath
        self.Workbook = None
        self.NameColumn = 1
        self.KeywordColumn = 3
        self.csvLock = threading.Lock()

    def AddVideo(self, vidName, metaDataList=None):
        # Get the template worksheet
        ws = self.Workbook.worksheets[0]

        # Always insert in the same place
        ws.insert_rows(3)

        # Write the name
        # TODO Make sure that the video file doens't already exist
        ws['A3'] = vidName

        # TODO Add options to add more data here
        if metaDataList == None:
            return
        else:
            self.AddMetadataTags(vidName, metaDataList)
            return

    def RemoveVideoEntry(self):
        # Remove an entry in the csv file
        return

    def CreateCSV(self, template, name):
        # Create a new black box csv file
        # Since we don't have a way to create the file, we copy and rename the Master.xlsx
        dir = 'C:\VMShared\GitRepos\PersonalProject\VideoTagger\Results'
        newName = os.path.join(dir, name + '.xlsx')
        # Copy the template
        shutil.copy(template, newName)

        # head_tail = os.path.split(template)

        # Create the path of the copied file
        # fileToRename = os.path.join(dir , head_tail[1])

        # Rename the template to our new Job specific name
        # os.rename(fileToRename, name)

        self.CSVPath = newName
        self.Workbook = self.OpenCSV(self.CSVPath)

        # Return the path of our new xlsx
        return self.CSVPath

    def OpenCSV(self, filePath=None):

        if filePath == None:
            filePath = self.CSVPath
        # Open CSV so that the class can edit the entries
        return pyxl.load_workbook(filePath)

    def SaveCSV(self, name=None):
        self.Workbook.save(self.CSVPath)

    def LockCSV(self):
        self.csvLock.acquire()
        return

    def UnlockCSV(self):
        self.csvLock.release()
        return

    def AddMetadataTags(self, fileName, tags):
        # Add a list tags (str type) to a video entry in the csv
        # Find the row for the video
        ws = self.Workbook.worksheets[0]
        for rw in range(2, ws.max_row+1):
            cellvalue = ws.cell(row=rw, column=self.NameColumn).value
            if cellvalue == fileName:
                # Found a row matching the fileName
                # Add the list of metadata to the tags column for this row
                tags_list = ListToString(tags)
                ws.cell(row=rw, column=self.KeywordColumn).value = tags_list
        return

    def FindRow(self, key, clmn):
        csv = self.OpenCSV()

        # Given a search key, look through each cell in the column and return the row numbers where it exists
        for row in range(0, 20):
            cellValue = csv.cell(column=clmn, row=row,
                                 value="{0}".format(get_column_letter(col)))
            if cellValue == key:
                return


def main():

    name = 'Testing' + str(time.time_ns())
    
    csvObj = BlackBoxCSVManager()
    csvObj.CreateCSV(csvTemplate, name)

    logging.basicConfig(level=logging.INFO,
                        format='%(levelname)s: %(message)s')

    print("Creating Amazon S3 bucket and uploading video.")
    s3_resource = boto3.resource('s3')
    bucket = s3_resource.create_bucket(
        Bucket=f'doc-example-bucket-rekognition-{time.time_ns()}',
        CreateBucketConfiguration={
            'LocationConstraint': s3_resource.meta.client.meta.region_name
        })

    paths = FindItemsInDirectory(targetDir, videoExtensions)
    
    # Start asyncronous upload and rekognition tasks
    CreateBlackBoxCSVWithRekog(paths, bucket, csvObj, False)

    print("Task Complete.")
    print("Deleting resources.")
    bucket.objects.delete()
    bucket.delete()
    logger.info("Deleted bucket %s.", bucket.name)
    print("All resources cleaned up.")
    print('-'*88)

if __name__ == '__main__':
    main()
